# -*- coding: utf-8 -*-
"""Заполняет log.xlsx историей из История тренировок.txt"""
import re
import openpyxl
from pathlib import Path

DESKTOP = Path(r"c:\Users\user\Desktop")
HISTORY_TXT = DESKTOP / "История тренировок.txt"
LOG_XLSX = DESKTOP / "log.xlsx"

# Паттерны
DATE_RE = re.compile(r"^(\d{4}\.\d{2}\.\d{2})\s*$")
SET_RE = re.compile(r"^(\d+(?:\.\d+)?)\s*кг\s*×\s*(\d+)\s*повторений\s*$")
REST_RE = re.compile(r"^отдых\s+(\d+(?:\.\d+)?)\s*м\s*$", re.IGNORECASE)


def parse_history(path: Path):
    """Парсит История тренировок.txt. Возвращает список словарей для строк Excel."""
    text = path.read_text(encoding="utf-8")
    lines = [line.strip() for line in text.splitlines()]
    rows = []
    current_date = None
    current_exercise = None
    rest_next = 0

    i = 0
    while i < len(lines):
        line = lines[i]
        if not line:
            i += 1
            continue

        # Дата: 2026.02.09
        if DATE_RE.match(line):
            current_date = line.replace(".", "-")  # 2026-02-09
            i += 1
            continue

        # Упражнение: "Сгибание ног лежа / Prone leg curl" или "Тяга гантелей на наклонной скамье 30°"
        # (не дата, не подход, не отдых, не суперсет; не строка типа "42м" или "Ноги • Плечи")
        if current_date and "кг" not in line and "отдых" not in line.lower() and not line.startswith("СУПЕРСЕТ"):
            if "/" in line or ("•" not in line and not re.match(r"^\d+м\s*$", line)):
                current_exercise = line
            i += 1
            continue

        # Подход: "27 кг × 15 повторений"
        set_m = SET_RE.match(line)
        if set_m and current_date and current_exercise:
            weight = float(set_m.group(1).replace(",", "."))
            reps = int(set_m.group(2))
            rest = rest_next
            rest_next = 0
            # Формат даты как в существующем xlsx: "2026.02.09, 12:00"
            dt = current_date.replace("-", ".") + ", 12:00"
            rows.append({
                "date": dt,
                "exercise_name": current_exercise,
                "input_weight": weight,
                "total_weight": weight,
                "reps": reps,
                "rest": rest,
            })
            i += 1
            continue

        # Отдых: "отдых 2м"
        rest_m = REST_RE.match(line)
        if rest_m:
            rest_next = float(rest_m.group(1).replace(",", "."))
            i += 1
            continue

        rest_next = 0
        i += 1

    return rows


def main():
    rows = parse_history(HISTORY_TXT)
    if not rows:
        print("Не найдено записей в", HISTORY_TXT)
        return

    wb = openpyxl.load_workbook(LOG_XLSX)
    sheet = wb.active
    header = ["Date", "Exercise_ID", "Exercise_Name_Calc", "Input_Weight", "Total_Weight", "Reps", "Rest"]

    # Заголовок уже есть в файле; ищем строку для вставки
    first_row_vals = [c.value for c in sheet[1]] if sheet.max_row >= 1 else []
    if sheet.max_row == 0 or first_row_vals != header:
        sheet.insert_rows(1)
        for col, h in enumerate(header, 1):
            sheet.cell(row=1, column=col, value=h)

    start_row = sheet.max_row + 1
    for r in rows:
        sheet.append([
            r["date"],
            "",  # Exercise_ID оставляем пустым
            r["exercise_name"],
            r["input_weight"],
            r["total_weight"],
            r["reps"],
            r["rest"],
        ])

    wb.save(LOG_XLSX)
    print(f"Добавлено записей: {len(rows)}. Всего строк в листе: {sheet.max_row}. Файл: {LOG_XLSX}")


if __name__ == "__main__":
    main()
